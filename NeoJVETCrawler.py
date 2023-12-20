#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#----------------------------------------------------------------------------
# Created By  : João Santos
# Created Date: 2023/12/13
# Updated Date: 2023/12/14
# version ='1.0.1'
#
# Description:
#     JVET Meetings crawler, fetches meeting files and all documents
#
# To Do:
#  - Smarter resume which would check already downloaded/extracted files
#  - Parallel fetch or extraction or fetch and extraction
# ---------------------------------------------------------------------------

__author__ = "João Santos"
__copyright__ = "Copyright 2023, João Santos"
__license__ = "GPL2"
__version__ = "1.0.1"
__maintainer__ = "João Santos"
__email__ = "joaompssantos@gmail.com"
__status__ = "Production"


import argparse
from bs4 import BeautifulSoup
import openpyxl
import os
import pandas
import shutil
from tabulate import tabulate
import urllib
import zipfile


### To Do
# Make a proper header for the file with copyright and what not


# Pause function for debug
def pause():
    programPause = input("\nPress the <ENTER> key to continue...")


# Function to deal with the input arguments
def getArgs():
    parser = argparse.ArgumentParser(description='Get and keep up to date the documentation and meeting notes of JVET')
    parser.add_argument('-v', '--verbose', dest='verbose', action='store_true', required=False, help='verbose mode to get extra information')
    parser.add_argument('-p', '--pause', dest='pause', action='store_true', required=False, help='pause on verbose')
    parser.add_argument('-s', '--nosavexls', dest='savexls', action='store_false', required=False, help='disable saving information as xls file')
    parser.add_argument('-r', '--rmzip', dest='rmzip', action='store_true', required=False, help='remove zip files after extraction')
    parser.add_argument('-f', '--force', dest='force', action='store_true', required=False, help='force to redo operations that would be skipped')
    parser.add_argument('-d', '--docsource', dest='docsource', nargs=1, type=str, required=False,
                        help='link to the page with the list of all JVET meetings (might not work if changed)',
                        default = 'https://www.jvet-experts.org/doc_end_user/all_meeting.php')
    parser.add_argument('-n', '--notesource', dest='notesource', nargs=1, type=str, required=False,
                        help='link to the page with the list of all JVET meeting notes (might not work if changed)',
                        default = 'https://www.itu.int/wftp3/av-arch/jvet-site/')
    parser.add_argument('-z', '--zipdir', dest='zipdir', nargs=1, type=str, required=False, help='directory to store the zip files',
                        default = 'zipfiles')

    requiredNamed = parser.add_argument_group('required arguments')
    requiredNamed.add_argument('-o', '--outputdir', dest='outputdir', type=str, required=True, help='directory to store the documents')

    return parser.parse_args()


# Save list to xls
def saveXlsFile(list, path):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Append each row in the table to the worksheet
    for row in list:
        ws.append(row)
    
    # Save the workbook
    wb.save(os.path.expanduser(path))


# Function to build a table of relevant information of all meetings from the meetings links (docs + notes)
def getAllMeetingsTable(args):
    # Links to the information and documents to fetch
    all_meetings_url = args.docsource
    notes_meetings_url = args.notesource

    # Get url source to be parsed by BeautifulSoup
    all_meetings_source = urllib.request.urlopen(all_meetings_url)
    # Read html from source
    all_meetings_soup = BeautifulSoup(all_meetings_source, 'lxml')
    # Table iterator
    all_meetings_table = all_meetings_soup.table.findAll('tr')
    # Links iterator with extra added so that it has the same number of elements as the table
    all_meetings_links = BeautifulSoup('<a href="Meeting Link"></a>', 'lxml').find_all('a') + all_meetings_soup.table.findAll('a')

    # Get url source to be parsed by BeautifulSoup
    notes_meetings_source = urllib.request.urlopen(notes_meetings_url)
    # Read html from source
    notes_meetings_soup = BeautifulSoup(notes_meetings_source, 'lxml')
    # Links iterator extracted to a list for convinience
    notes_meetings_links = ['Notes Link']
    
    # Loop the links to find those of an actual meeting (starting with /wftp3/av-arch/jvet-site/2)
    for links in reversed(notes_meetings_soup.findAll('a')):
        link = links.get('href')

        # Check if link conforms to expectation
        if link.startswith(urllib.parse.urlparse(notes_meetings_url).path + '2'):
            notes_meetings_links.append(f"{urllib.parse.urlparse(notes_meetings_url).scheme}://{urllib.parse.urlparse(notes_meetings_url).netloc}{link}")
    
    # Add an empty string to keep all lists the same size since the first meeting does not appear to have notes
    notes_meetings_links.append('')

    # Alocate meeting info table
    meeting_info_table = []

    # Check for first element
    first = 0

    # Loop table rows to extract information from both links
    for infos, mlinks, nlinks in zip(all_meetings_table, all_meetings_links, notes_meetings_links):
        # Get information from table row
        info = infos.find_all('td')
        info_row = [i.text for i in info]

        # Get corrependent link
        link = ''
        if first == 0:
            first = 1
        else:
            link = all_meetings_url.replace('all_meeting.php', '')

        # Append both links to row
        info_row.append(nlinks)
        info_row.append(link + mlinks.get('href'))

        # Append to table
        meeting_info_table.append(info_row)
    
    # Convert meetings numbers to int
    for row in meeting_info_table[1:]:
        row[0] = int(row[0])
    # Sort meetings by their number
    meeting_info_table[1:] = sorted(meeting_info_table[1:])

    # Print table
    if args.verbose:
        print('Global table with all the meetings info:\n')
        print(tabulate(meeting_info_table, headers='firstrow'))

        if args.pause:
            pause()
    
    # Save table to xls
    if args.savexls:
        saveXlsFile(meeting_info_table, os.path.join(args.outputdir, '#all_meetings_info.xlsx'))

    # Return table
    return meeting_info_table


# Function to try and fetch the zip url from the preview page
def fetchZipUrl(doc_number, prev_url):
    zip_link = None

    # Open prev_url with BeautifulSoup
    html_page = urllib.request.urlopen(prev_url)
    soup = BeautifulSoup(html_page, 'lxml')

    # Check all the urls and extract the last which corresponds with the doc_number
    for link in soup.findAll('a'):
        if doc_number in link:
            zip_link = link.get('href').replace('..', '')

    return zip_link


# Get table of meeting docs
def getDocsTable(args, meeting_url):
    # Get meeting raw table from meeting page
    meeting_raw_table = pandas.read_html(meeting_url, extract_links = 'all')[1]

    # Drop uneeded columns (namely: MPEG number, Created and First upload)
    meeting_raw_table.drop([1, 2, 3], axis=1, inplace=True)

    # Create the actual meeting table with all information
    meeting_table = [['JVET Number', 'Title', 'Zip', 'Authors', 'Last Uploaded']]

    # Loop raw table skipping first line (headers)
    for irow in range(1, meeting_raw_table.shape[0] - 1):
        # Check if exists and skip if it does not
        if isinstance(meeting_raw_table.iloc[irow, 4], float):
            continue
        # Check if withdrawn and skip
        elif meeting_raw_table.iloc[irow, 4][0].lower() == 'withdrawn':
            continue
        # Sometimes zip link does not exist but it does not seem to be withdrawn
        # This addresses that case
        elif meeting_raw_table.iloc[irow, 4][1] is None:
            preview_page_url = urllib.parse.urljoin(args.docsource.replace('all_meeting.php', ''),
                                                    meeting_raw_table.iloc[irow, 0][1])
            zip_url = fetchZipUrl(meeting_raw_table.iloc[irow, 0][0], preview_page_url)
            
            # If for some reason no link was found the current doc is skipped
            if zip_url is None:
                continue
        else:
            # Generate proper zip url
            zip_url = meeting_raw_table.iloc[irow, 4][1].replace('..', '')

        # Generate full zip url
        zip_url = urllib.parse.urljoin(args.docsource.replace('doc_end_user/all_meeting.php', ''), zip_url)

        curr_doc = [meeting_raw_table.iloc[irow, 0][0], # JVET Number
                    meeting_raw_table.iloc[irow, 2][0], # Title
                    zip_url,                            # Link to zip
                    meeting_raw_table.iloc[irow, 3][0], # Author list
                    meeting_raw_table.iloc[irow, 1][0]] # Last uploaded date

        meeting_table.append(curr_doc)

    return meeting_table


# Get notes and logistics links
def getNotesLinks(args, notes_url):
    notes_link = None
    logistics_link = None

    # Open prev_url with BeautifulSoup
    html_page = urllib.request.urlopen(notes_url)
    soup = BeautifulSoup(html_page, 'lxml')

    # Check all the urls and extract the last which corresponds with the notes or logistics tag
    for link in soup.findAll('a'):
        link = link.get('href')
        if 'notes' in link.lower():
            notes_link = urllib.parse.urljoin(args.notesource, '/'.join(link.split('/')[-2:]))
        if 'logistics' in link.lower():
            logistics_link = urllib.parse.urljoin(args.notesource, '/'.join(link.split('/')[-2:]))

    return [notes_link, logistics_link]


# Save meeting infos to xls
def saveMeetingInfosXlsFile(meeting_name, no_docs, docs_list, notes_links, path):
    # Create workbook and make active
    wb = openpyxl.Workbook()
    ws = wb.active

    # Add meeting title
    ws.append(['Meeting:', meeting_name])
    # Make it bold
    ws['A1'].font = openpyxl.styles.Font(bold = True)
    ws['B1'].font = openpyxl.styles.Font(bold = True)
    # Add link to folder
    ws['B1'].hyperlink = os.path.dirname(path)
    # Skip row
    ws.append([''])

    # Add links to notes and logistics
    ws.append(['Notes:', notes_links[0][0]])
    ws.append(['Logistics:', notes_links[1][0]])
    # Add links to local files
    ws['A3'].hyperlink = notes_links[0][1]
    ws['A4'].hyperlink = notes_links[1][1]
    # Add links to remote files
    ws['B3'].hyperlink = notes_links[0][0]
    ws['B4'].hyperlink = notes_links[1][0]
    # Skip row
    ws.append([''])

    # Add number of docs
    ws.append(['Number of docs:', no_docs])
    # Skip row
    ws.append([''])

    # Append headers to the worksheet
    ws.append(docs_list[0])

    # Add links to docs folders and to zip url
    for ix, doc in zip(range(len(docs_list) - 1), docs_list[1:]):
        # Append each row in the docs list to the worksheet
        ws.append(doc)
        # Target cell
        cell = f'A{9 + ix}'
        ws[cell].hyperlink = doc[0]

        # Target cell
        cell = f'C{9 + ix}'
        ws[cell].hyperlink = doc[2]
    
    # Save the workbook
    wb.save(os.path.expanduser(path))


# Function to collect all relevant information of a single meeting (docs + notes)
def getMeetingInfos(args, meeting_path, meeting_info):
    # Links to the information and documents to fetch
    meeting_url = meeting_info[-1]
    notes_url = meeting_info[-2]

    # Get table of meeting docs
    docs_table = getDocsTable(args, meeting_url)

    # Get number of docs in this meeting
    no_docs = len(docs_table) - 1

    if notes_url == '':
        notes_links = [None, None]
    else:
        notes_links = getNotesLinks(args, notes_url)

    # Print table
    if args.verbose:
        print(meeting_info)
        print(f'Infos for meeting {os.path.basename(meeting_path)} (number of docs: {no_docs}):')
        print(f'    Meeting output directory: {meeting_path}')
        print(f'    Meeting notes url: {notes_url}')

        print('')
        print(tabulate(docs_table, headers='firstrow'))

        if args.pause:
            pause()

    return no_docs, docs_table, notes_links


# Download notes and logistics files
def fetchNotesLogistics(notes_urls, meeting_folder):
    notes_file = ''
    logistics_file = ''

    if not notes_urls[0] is None:
        # Notes out file name
        notes_file = os.path.join(meeting_folder, urllib.parse.urlparse(notes_urls[0]).path.split('/')[-1])
        # Fetch file to notes_file
        urllib.request.urlretrieve(notes_urls[0], notes_file)

    if not notes_urls[1] is None:
        # Notes out file name
        logistics_file = os.path.join(meeting_folder, urllib.parse.urlparse(notes_urls[1]).path.split('/')[-1])
        # Fetch file to logistics_file
        urllib.request.urlretrieve(notes_urls[1], logistics_file)

    return [[notes_urls[0], notes_file], [notes_urls[1], logistics_file]]


# Download notes and logistics files
def fetchZipFiles(docs_table, zip_folder):
    # Create a list with the zip files location
    zip_files = []

    # Docs number
    no_docs = len(docs_table) - 1

    # Loop docs_table
    for zip_link, ix in zip(docs_table[1:], range(len(docs_table[1:]))):
        # zip out file name
        zip_file = os.path.join(zip_folder, urllib.parse.urlparse(zip_link[2]).path.split('/')[-1])
        
        print(f'            [{ix + 1:04} out of {no_docs:04}] Downloading {zip_link[0]} ...', end='')
        # Fetch file to notes_file
        urllib.request.urlretrieve(urllib.parse.quote(zip_link[2], safe=':/'), zip_file)
        # Print a message indicating that the extraction is complete
        print('    Done!')
        # Append file to list
        zip_files.append(zip_file)

    return zip_files


# Extract all meeting files
def extractZipFiles(args, docs_table, zip_files, meeting_folder):
    # Create an error list for files that can't be extracted
    errorlist = []

    # Docs number
    no_docs = len(docs_table) - 1

    # Loop docs_table
    for curr_doc, zip_file, ix in zip(docs_table[1:], zip_files, range(len(docs_table[1:]))):
        curr_doc = curr_doc[0]

        try:
            with zipfile.ZipFile(zip_file, 'r') as archive:
                print(f'            [{ix + 1:04} out of {no_docs:04}] Extracting {curr_doc} ...', end='')
                # Extract all contents of the zip file to a directory with the doc number name
                archive.extractall(path=os.path.join(meeting_folder, curr_doc))
                # Print a message indicating that the extraction is complete
                print('    Done!')
        except zipfile.BadZipfile:
            errorlist.append(f'{curr_doc}:    {zip_file}')
        
        # Remove zip files if option is set
        if args.rmzip == 'yes':
            os.remove(zip_file)
    
    return errorlist


# Parse meeting info table
def parseGlobalInfo(args, meeting_info_table):
    # Number of meetings
    no_meetings = len(meeting_info_table[1:])

    # Loop table meetings
    for meeting_row, ix in zip(meeting_info_table[1:], range(no_meetings)):
        # Meeting name YYYY_MM_L_CITY
        meeting_name = f"{meeting_row[2].split('-')[0]}_{meeting_row[2].split('-')[1]}_{meeting_row[4]}_{meeting_row[1].replace(' ', '_')}"
        # NR_L_CITY_YYYY_MM
        # meeting_name = f"{int(meeting_row[0]):03}_{meeting_row[4]}_{meeting_row[1].replace(' ', '_')}_{meeting_row[2].split('-')[0]}_{meeting_row[2].split('-')[1]}"))

        # Defines the folder name in the format: YYYY_MM_L_CITY
        meeting_folder = os.path.expanduser(os.path.join(args.outputdir, meeting_name))

        # Defines the name of the file where all the meeting information is stored
        meeting_file = os.path.join(meeting_folder, '#meeting_info.xlsx')

        # Skips current if the corresponding folder already exists
        # If the meeting informations file does not exist or the force option
        # is activated the current meeting operations are performed form scratch
        # The same happens for the last meeting
        if os.path.exists(meeting_folder):
            if args.force                           \
               or not os.path.exists(meeting_file)  \
               or meeting_row == meeting_info_table[-1]:
                shutil.rmtree(meeting_folder)
            else:
                continue
        
        print(f'    [{ix + 1:03} out of {no_meetings:03}] Working on meeting {meeting_name}...')
        
        # Create directory for meeting
        os.mkdir(meeting_folder)

        # Get current meeting table
        print('        Fetching meeting infos...')
        no_docs, docs_table, notes_links = getMeetingInfos(args, meeting_folder, meeting_row)
        print('        Meeting infos fetched!\n')

        # Download notes and logistics files
        print('        Fetching notes and logistics files...')
        notes_links = fetchNotesLogistics(notes_links, meeting_folder)
        print('        Files fetched!\n')

        # Download zip files
        print('        Fetching doc zip files...')
        os.mkdir(os.path.join(meeting_folder, args.zipdir))
        zip_files = fetchZipFiles(docs_table, os.path.join(meeting_folder, args.zipdir))
        print('        Zip files fetched!\n')

        # Unzip zip files
        print('        Extracting doc zip files...')
        error_list = extractZipFiles(args, docs_table, zip_files, meeting_folder)
        # Remove zip files if option is set
        if args.rmzip == 'yes':
            os.remove(os.path.join(meeting_folder, args.zipdir))
        print('        Zip files extracted!\n')

        # If there were errors during the extraction
        if len(error_list) > 0:
            # Error file path
            error_file = os.path.join(meeting_folder, '#extraction_error_list.txt')

            # Save list to file
            with open(error_file, 'w') as fp:
                for error in error_list:
                    fp.write(f'{error}')
            
            # Information messages
            print(f'{len(error_list)} file(s) could not be extracted.')
            print(f'A file with details was saved to: {error_file}.\n')
        
        # Save infos to xls
        if args.savexls:
            saveMeetingInfosXlsFile(os.path.basename(meeting_folder), no_docs, docs_table, notes_links,
                                    meeting_file)
        
        print(f'    [{ix + 1:03} out of {no_meetings:03}] Finished meeting {meeting_name}!\n')


# Defining main function 
def main():
    # Parse arguments
    args = getArgs()

    print('Fetching all JVET documents, please wait...\n')

    # Get all meetings table
    print('Compiling table with all meetings information...')
    meeting_info_table = getAllMeetingsTable(args)
    print('Table compiled!\n')

    # Parse the previous table information and download files
    print('Parsing all meetings...')
    parseGlobalInfo(args, meeting_info_table)
    print('Parsing completed!\n')

    print('All files fetched!')


# Call main function
if __name__=="__main__": 
    main() 